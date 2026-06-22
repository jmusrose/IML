import sys
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

import torch

INTRO_ROOT = Path(__file__).resolve().parents[1] / "Introduction_experiment"
sys.path.insert(0, str(INTRO_ROOT))
for name in list(sys.modules):
    if name == "AV_v1" or name.startswith("AV_v1."):
        del sys.modules[name]

from AV_v1.models.baseline import AVBaseline
from AV_v1.train_cremad import (
    compute_loss_gaps,
    freeze_encoder,
    plot_gap_history,
    write_gap_excel,
)


class IntroCremadExperimentTest(unittest.TestCase):
    def test_av_baseline_uses_independent_probe_heads(self):
        model = AVBaseline(num_classes=6, fusion_head="linear")
        model.eval()

        audio = torch.randn(2, 1, 64, 80)
        visual = torch.randn(2, 3, 2, 64, 64)

        with torch.no_grad():
            outputs = model.forward_with_modal_logits(audio, visual)

        self.assertEqual(outputs["logits"].shape, (2, 6))
        self.assertEqual(outputs["audio_logits"].shape, (2, 6))
        self.assertEqual(outputs["visual_logits"].shape, (2, 6))
        self.assertIsNot(model.audio_probe, model.classifier)
        self.assertIsNot(model.visual_probe, model.classifier)
        self.assertFalse(torch.allclose(outputs["audio_logits"], outputs["visual_logits"]))

    def test_all_requested_fusion_heads_forward(self):
        audio = torch.randn(2, 1, 64, 80)
        visual = torch.randn(2, 3, 2, 64, 64)

        for fusion_head in ("linear", "logit_fusion", "transencoder"):
            with self.subTest(fusion_head=fusion_head):
                model = AVBaseline(num_classes=6, fusion_head=fusion_head)
                model.eval()
                with torch.no_grad():
                    outputs = model.forward_with_modal_logits(audio, visual)
                self.assertEqual(outputs["logits"].shape, (2, 6))

    def test_loss_gaps_are_fusion_minus_each_probe(self):
        metrics = {
            "fusion_loss": 1.8,
            "audio_loss": 1.2,
            "visual_loss": 2.1,
        }

        gaps = compute_loss_gaps(metrics)

        self.assertAlmostEqual(gaps["fusion_minus_audio"], 0.6)
        self.assertAlmostEqual(gaps["fusion_minus_visual"], -0.3)

    def test_freeze_encoder_disables_only_requested_encoder(self):
        model = AVBaseline(num_classes=6)

        freeze_encoder(model, "audio")

        self.assertTrue(all(not param.requires_grad for param in model.audio_net.parameters()))
        self.assertTrue(any(param.requires_grad for param in model.visual_net.parameters()))
        self.assertTrue(any(param.requires_grad for param in model.classifier.parameters()))
        self.assertTrue(any(param.requires_grad for param in model.audio_probe.parameters()))

    def test_gap_excel_and_plot_are_written(self):
        history = [
            {
                "epoch": 1,
                "train": {
                    "fusion_minus_audio": 0.3,
                    "fusion_minus_visual": -0.2,
                },
            },
            {
                "epoch": 2,
                "train": {
                    "fusion_minus_audio": 0.1,
                    "fusion_minus_visual": 0.4,
                },
            },
        ]

        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            excel_path = root / "loss_gaps.xlsx"
            plot_path = root / "loss_gaps.png"

            write_gap_excel(history, excel_path)
            plot_gap_history(history, plot_path)

            self.assertTrue(excel_path.exists())
            self.assertTrue(plot_path.exists())
            self.assertGreater(plot_path.stat().st_size, 0)

            from openpyxl import load_workbook

            sheet = load_workbook(excel_path).active
            headers = [sheet.cell(row=1, column=column).value for column in range(1, 3)]
            values = [sheet.cell(row=2, column=column).value for column in range(1, 3)]
            self.assertEqual(headers, ["fusion_minus_audio", "fusion_minus_visual"])
            self.assertEqual(values, [0.3, -0.2])


if __name__ == "__main__":
    unittest.main()
